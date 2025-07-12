#!/usr/bin/env python3


# region Module Description And Imports.
"""Handles all interactions with Excel files for Thermo Tracker.

This module provides classes and methods to create, read, and update
Excel files for managing radiator data.
"""

from __future__ import annotations

import logging
import os
from typing import Any, Callable, cast

import openpyxl  # type: ignore # pylint: disable=import-error
import pyinputplus as pyip  # type: ignore # pylint: disable=import-error
from openpyxl.cell.cell import Cell, MergedCell  # pylint: disable=import-error
from tabulate import tabulate  # type: ignore # pylint: disable=import-error

import constants as cons
import utils
from excel_utils import ExcelSharedMethods

# Aliases for all the Enums of constants.py.
CKeys = cons.ConfigKeys
Files = cons.Files
Labels = cons.Labels
ManId = cons.ManagerIdentifiers
RegMeta = cons.RegistryExcelMeta
SCons = cons.SharedConstants
UserMsgs = cons.UserMessages


# endregion.


# region Manage Radiators Registry File.


class RegistryManager:
    """Manages the creation and updating of the radiators registry file.

    Orchestrates the creation of a new registry file or the updating of an
    existing one. It delegates tasks to the RegistryGenerator and
    RegistryUpdater classes.

    Attributes:
        config_data (dict[str, str]): Data loaded from or to be saved in the
            JSON config file.
        config_update (dict[str, bool]): Indicates if the configuration data
            needs to be updated.
        instances dict[ManId, object]: Instances of registry-related classes.
    """

    config_data: dict[str, Any]
    config_update: dict[str, bool]
    instances: dict[ManId, object]

    def __init__(self, config_data, config_update) -> None:
        """Initializes attributes and creates modules instances."""
        self.config_data = config_data  # pylint: disable=duplicate-code
        self.config_update = config_update
        self.instances = {
            ManId.REGISTRY_GENERATOR: RegistryGenerator(self.config_data, self.config_update),
            ManId.REGISTRY_UPDATER: None,  # Lazy initialization.
        }

    def update_registry_path(self) -> None:
        """Ensures the registry file path is stored in the JSON config file."""
        utils.update_config_path(
            self.config_data, self.config_update, CKeys.REGISTRY_FILE, Files.RADIATORS_REGISTRY
        )

    def setup_registry(self) -> None:
        """Runs the process of setting up the registry file.

        If the registry file does not exist, creates a new one. If the
        file exists, prompts the user to confirm whether to update the
        radiators data (ID, Coefficient).
        """
        registry_path = self.config_data[CKeys.REGISTRY_FILE]

        if not os.path.exists(registry_path):
            self._generate_registry(registry_path)
        else:
            self._handle_existing_registry(registry_path)

    def _generate_registry(self, registry_path: str) -> None:
        """Generates a registry file and saves it to the specified path.

        Args:
            registry_path: Path where the registry file will be saved.
        """
        generator = self._get_registry_generator()
        generator.generate_registry()
        generator.save_registry(registry_path)

    def _get_registry_generator(self) -> RegistryGenerator:
        """Gets the RegistryGenerator instance with casting for type safety.

        Returns:
            The RegistryGenerator instance.
        """
        return cast(RegistryGenerator, self.instances[ManId.REGISTRY_GENERATOR])

    def _handle_existing_registry(self, registry_path: str) -> None:
        """Allows the user to review and update the registry data as needed.

        Args:
            registry_path: Path to the existing radiators registry file.
        """
        updater = self._get_registry_updater(registry_path)
        updater.update_registry(registry_path)
        updater.save_changes()

    def _get_registry_updater(self, registry_path: str) -> RegistryUpdater:
        """Gets the RegistryUpdater instance with casting for type safety.

        Since the RegistryUpdater instance does not already exist, it is created
        and stored in the instances dictionary.

        Args:
            registry_path: Path to the radiators registry file.

        Returns:
            The RegistryUpdater instance.
        """
        self.instances[ManId.REGISTRY_UPDATER] = RegistryUpdater(registry_path)
        return cast(RegistryUpdater, self.instances[ManId.REGISTRY_UPDATER])


# endregion.


# region Generate Radiators Registry File.


class RegistryGenerator:
    """Handles the creation of a new radiators registry Excel file.

    Manages the creation of a registry file that contains static information
    about radiators (names, IDs, and coefficients).

    Attributes:
        config_data (dict[str, str]): Data loaded from or to be saved in the
            JSON config file.
        config_update (dict[str, bool]): Indicates if the configuration data
            needs to be updated.
        worksheet (Worksheet): The active worksheet in the workbook.
        excel_shared (ExcelSharedMethods): Instance of ExcelSharedMethods.
    """

    config_data: dict[str, Any]
    config_update: dict[str, bool]
    worksheet: openpyxl.worksheet.worksheet.Worksheet
    excel_shared: ExcelSharedMethods

    def __init__(self, config_data, config_update) -> None:
        """Initializes attributes and creates module instances."""
        self.config_data = config_data
        self.config_update = config_update
        workbook = openpyxl.Workbook()
        self.worksheet = cast(openpyxl.worksheet.worksheet.Worksheet, workbook.active)
        self.excel_shared = ExcelSharedMethods(workbook, self.worksheet)

    def generate_registry(self) -> None:
        """Runs the process that creates the registry file.

        Generates the registry file, adds headers and data based on user
        input, and customizes the Excel sheet.
        """
        logging.info("Creating the radiators registry file.")
        utils.display_user_info(UserMsgs.REGISTRY_GENERATION)

        self._populate_registry_worksheet()
        self.excel_shared.customize_worksheet(Labels.REGISTRY, RegMeta)

    def _populate_registry_worksheet(self) -> None:
        """Runs the process that populates the worksheet with radiator data.

        Appends headers and user-provided radiator data to the
        worksheet.
        """
        self.worksheet.append(RegMeta.HEADERS.value)

        for data in self._collect_radiator_data(self._prompt_radiators_owned()):
            self.worksheet.append(data)

    def _prompt_radiators_owned(self) -> int:
        """Prompts the user for the number of radiators to register.

        Returns:
            The number of radiators to register.
        """
        logging.info("Prompting user for the number of radiators.")

        while True:
            radiators_owned = pyip.inputNum("\nNumber of radiators to register? (Use digits)\n")
            confirm_num = f"Confirm {radiators_owned} radiators? (yes/no)\n"
            if pyip.inputYesNo(confirm_num) == SCons.AGREE:
                self.config_data[CKeys.RADIATORS_OWNED] = radiators_owned
                self.config_update[SCons.UPDATE] = True
                return radiators_owned

    def _collect_radiator_data(self, radiators_owned: int) -> list[list[object]]:
        """Collects radiator data for the specified number of radiators.

        Args:
            radiators_owned: The number of radiators to register.

        Returns:
            A list of [name, ID, coefficient] entries for each radiator.
        """
        logging.info("Prompting user to enter registry data (name, ID, coefficient).")

        registry_data = []

        for radiator_num in range(1, radiators_owned + 1):

            while True:
                name, radiator_id, coefficient = self._prompt_radiator_entry(radiator_num)
                if self._confirm_radiator_entry(radiator_num, name, radiator_id, coefficient):
                    registry_data.append([name, radiator_id, coefficient])
                    break

        return registry_data

    def _prompt_radiator_entry(self, radiator_num: int) -> tuple[str, int, float]:
        """Prompts the user to input name, ID, and coefficient for a radiator.

        Args:
            radiator_num: The index of the radiator being prompted.

        Returns:
            A tuple containing the radiator name, ID, and coefficient.
        """
        methods = [pyip.inputStr, pyip.inputNum, pyip.inputFloat]

        return tuple(
            method(f"Radiator {radiator_num} - {header}: ")
            for method, header in zip(methods, RegMeta.HEADERS.value)
        )

    def _confirm_radiator_entry(
        self, radiator_num: int, name: str, radiator_id: int, coefficient: float
    ) -> bool:
        """Asks the user to confirm the entered radiator data.

        Args:
            radiator_num: The index of the radiator being confirmed.
            name: The radiator's name.
            radiator_id: The radiator's ID.
            coefficient: The radiator's coefficient.

        Returns:
            True if the user confirms the data; False otherwise.
        """
        print(f"\nReview and confirm the data for Radiator {radiator_num}:\n")
        for header, value in zip(RegMeta.HEADERS.value, [name, radiator_id, coefficient]):
            print(f"- {header}: {value}")

        radiator_confirmation = f"\nConfirm Radiator {radiator_num}? (yes/no)\n"
        return pyip.inputYesNo(radiator_confirmation) == SCons.AGREE

    def save_registry(self, registry_path: str) -> None:
        """Saves the registry workbook to the specified file path.

        Args:
            registry_path: Path where the registry file will be saved.
        """
        self.excel_shared.save_workbook(registry_path, Labels.REGISTRY)


# endregion.


# region Update Radiators Registry File.


class RegistryUpdater:
    """Handles the updating of an existing radiators registry Excel file.

    Allows the user to review and optionally edit existing radiator entries
    (ID, and Coefficient) in the registry file.

    Attributes:
        registry_path (str): Path to the radiators registry file.
        worksheet (Worksheet): The active worksheet in the workbook.
        excel_shared (ExcelSharedMethods): Instance of ExcelSharedMethods.
    """

    registry_path: str
    worksheet: openpyxl.worksheet.worksheet.Worksheet
    excel_shared: ExcelSharedMethods

    def __init__(self, registry_path) -> None:
        """Initializes attributes and creates module instances."""
        self.registry_path = registry_path
        workbook = openpyxl.load_workbook(self.registry_path)
        self.worksheet = cast(openpyxl.worksheet.worksheet.Worksheet, workbook.active)
        self.excel_shared = ExcelSharedMethods(workbook, self.worksheet)

    def update_registry(self, registry_path: str) -> None:
        """Allows the user to review and update the registry if desired.

        Args:
            registry_path: Path to the radiators registry file.
        """
        logging.info("Asking user to review registry file for potential updates.")
        table_string = self.get_radiator_registry_table()

        if self._prompt_update_registry(registry_path, table_string):
            utils.display_user_info(UserMsgs.REGISTRY_UPDATE_DESCRIPTION)
            logging.info("The user decided to update the registry file data.")

            # User can only update ID and Coefficient for each radiator.
            # Radiator names are treated as unique identifiers and are not editable.
            # Allowing name changes would break the mapping between registry data
            # and usage data, potentially causing data integrity issues.
            for row in self.worksheet.iter_rows(min_row=RegMeta.FIRST_DATA_ROW.value):
                self._recap_radiator_data(row)
                self._prompt_radiator_update(row)

    def _prompt_update_registry(self, registry_path: str, table_string: str) -> bool:
        """Prompts the user to confirm whether to update the registry.

        Args:
            registry_path: Path to the radiators registry file.
            table_string: Formatted table with the registry data to display.

        Returns:
            True if the user agrees to update the registry; False otherwise.
        """
        update_prompt = UserMsgs.REGISTRY_UPDATE_RECAP.format(
            registry_path=registry_path, registry_table=table_string
        )
        return pyip.inputYesNo(update_prompt) == SCons.AGREE

    def get_radiator_registry_table(self) -> str:
        """Provides a formatted table of all radiators registry data.

        Returns:
            The formatted table.
        """
        data = [
            list(row)
            for row in self.worksheet.iter_rows(
                min_row=RegMeta.FIRST_DATA_ROW.value, values_only=True
            )
        ]
        headers = RegMeta.HEADERS.value
        return tabulate(data, headers=headers, tablefmt="github")

    def _recap_radiator_data(self, row: tuple[Cell | MergedCell, ...]) -> None:
        """Recaps the data (ID, Coefficient) for a specific radiator.

        Args:
            row: The list of cell objects representing the radiator's data.
        """
        radiator_name = row[RegMeta.RAD_NAME_ROW.value].value
        print(f"\nRecap for Radiator {radiator_name}:")
        self._handle_editable_fields(row, lambda header, cell: print(f"- {header}: {cell.value}"))

    def _handle_editable_fields(
        self, row: tuple[Cell | MergedCell, ...], callback: Callable[[str, Cell | MergedCell], None]
    ) -> None:
        """Applies a callback to editable fields (ID, Coefficient) in a row.

        Args:
            row: The list of cell objects representing the radiator's data.
            callback: Function to apply to each editable field.
        """
        headers = RegMeta.HEADERS.value[1:]
        cells = row[1:]
        for header, cell in zip(headers, cells):
            callback(header, cell)

    def _prompt_radiator_update(self, row: tuple[Cell | MergedCell, ...]) -> None:
        """Prompts the user to update the data for a specific radiator.

        Args:
            row: The list of cell objects representing the radiator's data.
        """
        if pyip.inputYesNo("\nUpdate any of these data? (yes/no)\n") == SCons.AGREE:
            self._handle_editable_fields(
                row, lambda header, cell: self._prompt_header_update(row, header, cell)
            )

    def _prompt_header_update(
        self, row: tuple[Cell | MergedCell, ...], header: str, cell: Cell | MergedCell
    ) -> None:
        """Prompts the user if they want to update a specific radiator header.

        Args:
            row: The list of cell objects representing the radiator's data.
            header: The header name of the column being modified.
            cell: The cell object to be modified.
        """
        radiator_name = row[RegMeta.RAD_NAME_ROW.value].value
        print(f"\n{radiator_name} current value for '{header}': {cell.value}")

        update_successful = False

        while not update_successful:
            if pyip.inputYesNo(f"Update '{header}'? (yes/no)\n") == SCons.AGREE:
                new_value = self._prompt_cell_value(header, cell)
                update_successful = self._confirm_to_update(cell, new_value)
            else:
                # If the user doesn't want to modify, skip to the next header.
                update_successful = True

    def _prompt_cell_value(self, header: str, cell: Cell | MergedCell) -> int | float:
        """Prompts the user to enter a new value for a specific cell.

        Handles the user input based on the type of the cell.

        Args:
            header: The header name of the column being updated.
            cell: The cell object to be updated.

        Returns:
            The new value for the cell, based on its type.
        """
        update_prompt = f"Enter a new value for '{header}':\n"

        match cell.value:
            case int():
                return pyip.inputNum(update_prompt)
            case float():
                return pyip.inputFloat(update_prompt)
            case _:
                coord = cell.coordinate
                logging.warning("Unexpected cell type: (%s) in %s", type(cell.value), coord)
                print(f"Problem with cell type: {coord}. Please check manually. No changes made.")
                return cast(Any, cell.value)

    def _confirm_to_update(self, cell: Cell | MergedCell, new_value: str | int | float) -> bool:
        """Asks the user to confirm the change before updating the cell.

        Args:
            cell: The cell object to be updated.
            new_value: The new value for the cell.

        Returns:
            True if the new_value was successful; False otherwise.
        """
        print(f"\nOld value: {cell.value} | New value: {new_value}")

        if pyip.inputYesNo("Confirm modification? (yes/no)\n") == SCons.AGREE:
            self.worksheet[cell.coordinate].value = new_value
            return True
        return False

    def save_changes(self) -> None:
        """Saves all changes made to the radiators registry to its path."""
        self.excel_shared.save_workbook(self.registry_path, Labels.REGISTRY)


# endregion.
