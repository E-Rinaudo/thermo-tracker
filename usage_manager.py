#!/usr/bin/env python3


# region Module Description And Imports.
"""Manages the configuration and creation of the usage file."""

from __future__ import annotations

import datetime
import logging
import os
import subprocess
import sys
from typing import Any, Callable, Iterable, cast

import openpyxl  # type: ignore # pylint: disable=import-error
import pyinputplus as pyip  # type: ignore # pylint: disable=import-error
from openpyxl.comments import Comment  # pylint: disable=import-error
from openpyxl.styles import Alignment  # pylint: disable=import-error

import constants as cons
import utils
from excel_utils import ExcelSharedMethods

# Aliases for all the Enums of constants.py.
Files = cons.Files
Folds = cons.Folders
CKeys = cons.ConfigKeys
UserMsgs = cons.UserMessages
ManId = cons.ManagerIdentifiers
SCons = cons.SharedConstants
SNums = cons.SharedNumbers
DForm = cons.DateFormats
RegexP = cons.RegexPatterns
Labels = cons.Labels
UsMeta = cons.UsageExcelMeta
OsPlatforms = cons.OsPlatforms

# endregion.


# region Manage Radiators Usage File.


class UsageManager:
    """Main controller for managing the usage file.

    Orchestrates the configuration and creation of the usage file. It delegates
    configuration tasks to UsageConfigurator and file creation tasks to
    UsageGenerator.

    Attributes:
        config_data (dict[str, str]): Data loaded from or to be saved in the
            JSON config file.
        config_update (dict[str, bool]): Indicates if the configuration data
            needs to be updated.
        instances dict[ManId, object]: Holds instances of usage-related classes.
    """

    config_data: dict[str, Any]
    config_update: dict[str, bool]
    instances: dict[ManId, object]

    def __init__(self, config_data, config_update) -> None:
        """Initializes attributes and creates modules instances."""
        self.config_data = config_data
        self.config_update = config_update
        self.instances = {
            ManId.USAGE_CONFIG: UsageConfigurator(self.config_data, self.config_update),
            ManId.USAGE_GENERATOR: UsageGenerator(self.config_data, self.config_update),
        }

    def update_usage_path(self) -> None:
        """Ensures the usage folder path is stored in the JSON config file."""
        utils.update_config_path(
            self.config_data, self.config_update, CKeys.USAGE_FOLDER_PATH, Folds.USAGE_FOLDER
        )

    def setup_usage(self) -> None:
        """Orchestrates the configuration and generation of the usage file.

        Handles both the configuration process, first run or updates,
        and the creation or updating of the usage Excel file.
        """
        self._configure_usage()
        self._generate_usage()

    def _configure_usage(self) -> None:
        """Handles the configuration process for the usage file.

        Prompts the user to set up or update configuration settings as
        needed, depending on whether this is the first run or an update.
        """
        configurator = self._get_usage_configurator()

        # Check only if the folder exists; the file is created with it.
        if not os.path.exists(self.config_data[CKeys.USAGE_FOLDER_PATH.value]):
            configurator.usage_config_setup()
        else:
            if configurator.prompt_change_usage_config():
                configurator.prompt_usage_config_updates()

    def _get_usage_configurator(self) -> UsageConfigurator:
        """Gets the UsageConfigurator instance with casting for type safety.

        Returns:
            The UsageConfigurator instance.
        """
        return cast(UsageConfigurator, self.instances[ManId.USAGE_CONFIG])

    def _generate_usage(self) -> None:
        """Creates or updates the usage file based on current configuration.

        Loads or creates the file, populates it with data, and saves the
        results.
        """
        generator = self._get_usage_generator()
        generator.update_file_path()
        generator.manage_usage()
        generator.save_usage()
        generator.open_usage()

    def _get_usage_generator(self) -> UsageGenerator:
        """Gets the UsageGenerator instance with casting for type safety.

        Returns:
            The UsageGenerator instance.
        """
        return cast(UsageGenerator, self.instances[ManId.USAGE_GENERATOR])


# endregion.


# region Radiators Usage Configurator.


class UsageConfigurator:
    """Handles the configuration of the usage file.

    Collects and manages configuration settings for the usage file, such as
    the filename, date format, and date input mode.

    Attributes:
        config_data (dict[str, str]): Data loaded from or to be saved in the
            JSON config file.
        config_update (dict[str, bool]): Indicates if the configuration data
            needs to be updated.
    """

    config_data: dict[str, Any]
    config_update: dict[str, bool]

    def __init__(self, config_data, config_update) -> None:
        """Initializes attributes."""
        self.config_data = config_data
        self.config_update = config_update

    def usage_config_setup(self) -> None:
        """Collects config settings for the usage file on the first run."""
        logging.info("Starting first-time setup for the configurations of the usage file.")
        utils.display_user_info(UserMsgs.USAGE_GENERATION.value)

        self._prompt_usage_name()
        self._prompt_date_format()
        self._prompt_date_input_mode()
        self.config_update[SCons.UPDATE.value] = True

    def _prompt_usage_name(self) -> None:
        """Prompts user to provide two years to use in the usage file name."""
        start_year = datetime.datetime.now().year
        end_year = start_year + SNums.ONE.value
        years = UserMsgs.YEARS_FILENAME_ENTRY.value.format(
            usage_file=Files.DEFAULT_USAGE_NAME.value, start_year=start_year, end_year=end_year
        )
        self._prompt_config_option(
            years, RegexP.YEARS.value, CKeys.USAGE_NAME.value, embed_years=True
        )

    def _prompt_config_option(
        self, prompt: str, regex: str, config_key: str, embed_years: bool = False
    ) -> None:
        """Collects a config option (years, date format, or date input mode).

        Args:
            prompt: The message displayed to the user.
            regex: Regex pattern to validate input.
            config_key: Key that stores the provided value in the JSON config.
            embed_years: If True, embeds the input years into a filename format.
        """
        logging.info("Prompting config option for: %s", config_key)

        while True:
            entry = pyip.inputStr(prompt, allowRegexes=[regex], blockRegexes=[RegexP.BLOCK.value])

            if embed_years:
                entry = Files.DEFAULT_USAGE_NAME.value.replace(Files.YEARS_PLACEHOLDER.value, entry)

            if pyip.inputYesNo(f"Confirm: {entry}? (yes/no)\n") == SCons.AGREE.value:
                self.config_data[config_key] = entry
                break

    def _prompt_date_format(self) -> None:
        """Prompts user to choose European or American date formats."""
        date = UserMsgs.PROMPT_DATE_FORMAT.value
        self._prompt_config_option(date, RegexP.DATE_REGEX.value, CKeys.DATE_FORMAT.value)

    def _prompt_date_input_mode(self) -> None:
        """Prompts user to choose how to enter dates into the Excel file."""
        mode = UserMsgs.PROMPT_DATE_INPUT_MODE.value
        self._prompt_config_option(mode, RegexP.DATE_INPUT_REGEX.value, CKeys.DATE_INPUT_MODE.value)

    def prompt_change_usage_config(self) -> bool:
        """Prompts the user whether to modify the usage configurations.

        Returns:
            True if the user wants to modify the configuration; False otherwise.
        """
        logging.info("Prompting user whether to modify existing usage file configurations.")
        modify = UserMsgs.CHANGE_USAGE_CONFIG.value.format(
            file_name=self.config_data[CKeys.USAGE_NAME.value],
            date_format=self.config_data[CKeys.DATE_FORMAT.value],
            date_input=self.config_data[CKeys.DATE_INPUT_MODE.value],
        )

        return pyip.inputYesNo(modify) == SCons.AGREE.value

    def prompt_usage_config_updates(self) -> None:
        """Prompts the user to update existing usage file configurations."""
        logging.info("User decided to modify the usage file configurations.")
        config_items = [
            ("usage filename", CKeys.USAGE_NAME.value, self._prompt_usage_name),
            ("date format", CKeys.DATE_FORMAT.value, self._prompt_date_format),
            ("date input mode", CKeys.DATE_INPUT_MODE.value, self._prompt_date_input_mode),
        ]

        for label, key, update_method in config_items:
            prompt = f"Change {label}? (yes/no)\nCurrent: {self.config_data[key]}\n"
            if pyip.inputYesNo(prompt) == SCons.AGREE.value:
                self._make_config_change(update_method)

    def _make_config_change(self, update_method: Callable[[], None]) -> None:
        """Executes a config update and marks the JSON as needing an update.

        Args:
            update_method: The method to execute the configuration update.
        """
        update_method()
        self.config_update[SCons.UPDATE.value] = True


# endregion.


# region Generate Radiators Usage File.


class UsageGenerator:
    """Handles the creation and population of the usage file.

    Attributes:
        config_data (dict[str, str]): Data loaded from or to be saved in the
            JSON config file.
        config_update (dict[str, bool]): Indicates if the configuration data
            needs to be updated.
        worksheet (Worksheet): The active worksheet in the workbook.
        excel_shared (ExcelSharedMethods): Instance of ExcelSharedMethods.
        file_path (str): The path of the usage Excel file.
    """

    config_data: dict[str, Any]
    config_update: dict[str, bool]
    worksheet: openpyxl.worksheet.worksheet.Worksheet
    excel_shared: ExcelSharedMethods
    file_path: str

    def __init__(self, config_data, config_update) -> None:
        """Initializes attributes."""
        self.config_data = config_data
        self.config_update = config_update
        self.worksheet = cast(openpyxl.worksheet.worksheet.Worksheet, None)  # Lazy initialization.
        self.excel_shared = cast(ExcelSharedMethods, None)  # Lazy initialization.
        self.file_path = SCons.EMPTY_STR.value  # Lazy initialization.

    def update_file_path(self) -> None:
        """Updates the file path based on the current configuration data.

        Retrieves the usage folder path and file name from the
        configuration data and assigns it to the file_path attribute.
        """
        self.file_path = os.path.join(
            self.config_data[CKeys.USAGE_FOLDER_PATH.value],
            self.config_data[CKeys.USAGE_NAME.value],
        )

    def manage_usage(self) -> None:
        """Creates or loads the usage Excel file and populates it.

        Ensures the usage folder exists, creates or loads the workbook,
        and populates the file with headers and data.
        """
        self._ensure_usage_folder()
        self._setup_workbook()

        logging.info("Starting population of the usage file.")
        self._initialize_usage_sheet()
        self._populate_usage()

    def _ensure_usage_folder(self) -> None:
        """Creates the usage Excel files folder if it doesn't exist."""
        folder_path = self.config_data[CKeys.USAGE_FOLDER_PATH.value]

        if not os.path.exists(folder_path):
            logging.info("Creating usage folder.")
            os.makedirs(folder_path, exist_ok=True)

    def _setup_workbook(self) -> None:
        """Creates a new workbook or load an existing one.

        Sets set up worksheet and Excel shared methods helper.
        """
        if not os.path.exists(self.file_path):
            workbook = openpyxl.Workbook()
        else:
            workbook = openpyxl.load_workbook(self.file_path)

        self.worksheet = cast(openpyxl.worksheet.worksheet.Worksheet, workbook.active)
        self.excel_shared = ExcelSharedMethods(workbook, self.worksheet)

    def _initialize_usage_sheet(self) -> None:
        """Customizes worksheet and adds headers if this is a new usage file.

        It is only executed on the first run, when the usage file does
        not yet exist. On subsequent runs, only new data is appended to
        the file.
        """
        if not os.path.exists(self.file_path):
            self.excel_shared.customize_worksheet(Labels.USAGE.value, UsMeta)
            self.worksheet.append(UsMeta.HEADERS.value)

    def _populate_usage(self) -> None:
        """Populates the usage file with new data for the current session."""
        self._update_start_rows()
        self._fill_registry_data()
        self._display_data_entry_intro()
        self._fill_dates()
        self._manage_valve_settings()
        self._get_raw_readings()
        self._get_actual_values()
        self._get_total()
        self._get_notes()
        self._add_blank_lines()

    def _update_start_rows(self) -> None:
        """Updates and stores the current and previous data block start rows.

        Sets the current run's start row and saves the previous run's
        start row in the JSON config for later reference
        """
        # start_row and last_start_row are always recalculated and set for each
        # new file or session. Any old values in the config are overwritten, so
        # there is no risk of carrying over incorrect row indices from previous
        # files. No explicit config cleanup is needed.
        start_row = self.worksheet.max_row + SNums.ONE.value
        last_start_row = self.config_data.get(CKeys.START_ROW.value, start_row)

        self.config_data[CKeys.START_ROW.value] = start_row
        self.config_data[CKeys.LAST_START_ROW.value] = last_start_row
        self.config_update[SCons.UPDATE.value] = True

    def _fill_registry_data(self) -> None:
        """Appends radiator data (Name, ID, Coefficient) from registry file.

        Loads the registry worksheet and appends each radiator's data to
        the usage worksheet, starting from the second column.
        """
        registry_worksheet = self._get_registry_worksheet()

        logging.info("Filling usage file with registry data.")
        for row in registry_worksheet.iter_rows(min_row=SNums.TWO.value, values_only=True):
            empty_row = self.worksheet.max_row + SNums.ONE.value

            for col_offset, value in enumerate(row, start=SNums.TWO.value):
                self.worksheet.cell(row=empty_row, column=col_offset, value=value)

    def _get_registry_worksheet(self) -> openpyxl.worksheet.worksheet.Worksheet:
        """Loads the registry Excel file worksheet.

        Returns:
            The registry worksheet.
        """
        registry_workbook = openpyxl.load_workbook(self.config_data[CKeys.REGISTRY_FILE.value])
        return cast(openpyxl.worksheet.worksheet.Worksheet, registry_workbook.active)

    def _display_data_entry_intro(self) -> None:
        """Explains how the  data entry process for the usage file works."""
        utils.display_user_info(UserMsgs.USAGE_DATA_ENTRY_INTRO.value)

    def _fill_dates(self) -> None:
        """Fills the date column for each radiator row and for the total row.

        Inserts the same date (either automatically or manually
        obtained) for each radiator row that was just appended to the
        worksheet, plus one additional date entry for the total row that
        follows the radiator data block.
        """
        if self.config_data[CKeys.DATE_INPUT_MODE.value] == SCons.DATE_AUTO.value:
            date = self._get_date_automatically()
        else:
            date = self._get_date_manually()

        radiators_owned, start_row = self._get_usage_block_info()
        self._fill_col(
            rows=range(start_row, radiators_owned + start_row + SNums.ONE.value),
            get_value=lambda row: date,
            col=SNums.ONE.value,
        )

    def _get_date_automatically(self) -> str:
        """Provides today's date in European or American format.

        Returns:
            The formatted date string.
        """
        return self._format_date(datetime.datetime.now())

    def _format_date(self, date: datetime.date) -> str:
        """Formats a date object according to the configured date format.

        Args:
            date: The date to format.

        Returns:
            The formatted date string.
        """
        if self.config_data[CKeys.DATE_FORMAT.value] == SCons.EUROPEAN.value:
            return date.strftime(DForm.EU_FORMAT.value)

        return date.strftime(DForm.US_FORMAT.value)

    def _get_date_manually(self) -> str:
        """Prompts the user for a date and formats it as configured.

        Returns:
            The formatted date string.
        """
        logging.info("Prompting user for manual date entry.")
        print(UserMsgs.ENTER_MANUAL_DATE.value)

        while True:
            day, month, year = self._prompt_date()

            if not (date_obj := self._check_date(year, month, day)):
                continue

            if date := self._confirm_date(self._format_date(date_obj)):
                return date

    def _prompt_date(self) -> tuple[int, int, int]:
        """Prompts the user for the day, month, and year.

        Returns:
            A tuple with the three parts of the date.
        """
        day = pyip.inputInt(
            f"Day ({SNums.ONE.value} - {SNums.THIRTY_ONE.value}): ",
            min=SNums.ONE.value,
            max=SNums.THIRTY_ONE.value,
        )
        month = pyip.inputInt(
            f"Month ({SNums.ONE.value} - {SNums.TWELVE.value}): ",
            min=SNums.ONE.value,
            max=SNums.TWELVE.value,
        )
        year = pyip.inputInt(f"Year (e.g., {datetime.datetime.now().year}): ")
        return day, month, year

    def _check_date(self, year: int, month: int, day: int) -> datetime.date | None:
        """Validates and constructs a date object from year, month, and day.

        Args:
            year: The year component of the date.
            month: The month component of the date.
            day: The day component of the date.

        Returns:
            The date object if valid; None otherwise.
        """
        try:
            return datetime.date(year, month, day)
        except ValueError as err:
            logging.info("User entered an invalid date: %s.", err)
            print(f"❌ Invalid date: ({err}).\nPlease provide a correct one.")
            return None

    def _confirm_date(self, date: str) -> str | None:
        """Prompts the user for date confirmation.

        Args:
            date: The formatted date to confirm.

        Returns:
            The date if confirmed; None otherwise.
        """
        print(f"The date you provided is: {date}.")

        if pyip.inputYesNo("Confirm this date? (yes/no)\n") == SCons.AGREE.value:
            return date

        return None

    def _get_usage_block_info(self) -> tuple[int, int]:
        """Gets the number of radiators and the current data block start row.

        Returns:
            A tuple with the radiators owned and the starting row for the
                current run.
        """
        radiators_owned = utils.get_radiators_owned(self.config_data)
        start_row = self.config_data[CKeys.START_ROW.value]
        return radiators_owned, start_row

    def _fill_col(
        self, rows: Iterable[int], get_value: Callable[[int], int | float | str], col: int
    ) -> None:
        """Fills a column for the specified rows with the appropriate data.

        Args:
            rows: An iterable of row indices to fill.
            get_value: Takes the row index and returns the value to write.
            col: The column index to fill.
        """
        for row in rows:
            value = get_value(row)
            self.worksheet.cell(row=row, column=col, value=value)

    def _manage_valve_settings(self) -> None:
        """Manages the process of entering valve settings for each radiator.

        If the usage file does not exist, prompts the user to enter
        valve settings for each radiator. Otherwise, recaps current
        valve settings and allows the user to modify them.
        """
        if not os.path.exists(self.file_path):
            self._enter_initial_valve_settings()
            self._add_valve_setting_note()
        else:
            self._update_valve_setting()

    def _enter_initial_valve_settings(self) -> None:
        """Prompts the user to enter valve settings for each radiator.

        After entry, the user is shown a recap of all valve settings and
        can make corrections.
        """
        utils.display_user_info(UserMsgs.VALVE_SETTING_ENTRY.value)
        logging.info("Prompting user for valve setting.")

        radiators_owned, start_row = self._get_usage_block_info()
        self._fill_col(
            rows=range(start_row, start_row + radiators_owned),
            get_value=lambda row: pyip.inputStr(
                f"Valve Setting for {self.worksheet.cell(row=row, column=SNums.TWO.value).value}: ",
            ),
            col=SNums.EIGHT.value,
        )

        self._recap_and_edit_column(
            "valve setting", SNums.EIGHT.value, self._enter_initial_valve_settings
        )

    def _recap_and_edit_column(self, label: str, col: int, method: Callable[[], None]) -> None:
        """Recaps and optionally allows modification of recently entered data.

        Displays each entered value (strictly valve setting or raw reading)
        for user review. If the user wants to make changes, the appropriate
        entry method is called again.

        Args:
            label: The label for the data being recapped
                ('valve setting', 'raw reading').
            col: The worksheet column index of the data to recap.
            method: The method to call if the user chooses to modify the data.
        """
        print(f"\nReview the {label}s you just entered. You can make corrections if needed:\n")

        radiators_owned, start_row = self._get_usage_block_info()
        for row in range(start_row, start_row + radiators_owned):
            radiator_name = self.worksheet.cell(row=row, column=SNums.TWO.value).value
            setting = self.worksheet.cell(row=row, column=col).value
            print(f"- {radiator_name} {label}: {setting}")

        change_prompt = f"\nWould you like to modify any of these {label}s? (yes/no)\n"
        if pyip.inputYesNo(change_prompt) == SCons.AGREE.value:
            method()
        else:
            print(f"✅ Great! Your {label}s have been saved.")

    def _add_valve_setting_note(self) -> None:
        """Adds a note to the valve setting column header.

        The note informs users that Excel may show a warning for mixed
        data types, but this is intentional and not an error.
        """
        note = Comment(
            "NOTE:\nNumbers saved as text. Excel warning expected.",
            "Thermo Tracker",
            height=SNums.ONE_HUNDRED_FIFTY.value,
            width=SNums.TWO_HUNDRED.value,
        )
        self.worksheet.cell(row=SNums.ONE.value, column=SNums.EIGHT.value).comment = note

    def _update_valve_setting(self) -> None:
        """Runs the sequence to update the valve settings for the radiators."""
        utils.display_user_info(UserMsgs.VALVE_SETTING_UPDATE.value)
        logging.info("Reviewing valve settings with the user.")

        radiators_owned, start_row = self._get_usage_block_info()
        last_start_row = self.config_data[CKeys.LAST_START_ROW.value]

        for i in range(radiators_owned):
            old_row = last_start_row + i
            new_row = start_row + i
            self._process_valve_setting_update(old_row, new_row)

    def _process_valve_setting_update(self, old_row: int, new_row: int) -> None:
        """Handles the update process for a single radiator's valve setting.

        Args:
            old_row: The row index for the previous run's value.
            new_row: The row index for the current run's value.
        """
        radiator_name, current_valve_setting = self._get_valve_info(old_row)

        while True:
            if new_valve_setting := self._prompt_valve_update(radiator_name, current_valve_setting):
                if self._confirm_valve_update(current_valve_setting, new_valve_setting, new_row):
                    break
            else:
                self._keep_existing_valve_setting(new_row, current_valve_setting)
                break

    def _get_valve_info(self, old_row: int) -> tuple[str, str]:
        """Gets the radiator name and its valve setting for a given row.

        Args:
            old_row: The worksheet row index from the previous run.

        Returns:
            A tuple containing the radiator name and its valve setting.
        """
        radiator_name = self.worksheet.cell(row=old_row, column=SNums.TWO.value).value
        current_valve_setting = self.worksheet.cell(row=old_row, column=SNums.EIGHT.value).value

        return cast(str, radiator_name), cast(str, current_valve_setting)

    def _prompt_valve_update(self, radiator_name: str, current_valve_setting: str) -> str | None:
        """Prompts the user to update the valve setting for a radiator.

        Args:
            radiator_name: The name of the radiator.
            current_valve_setting: The current valve setting to be modified.

        Returns:
            The new valve setting if the user chooses to update; None otherwise.
        """
        change_prompt = UserMsgs.PROMPT_VALVE_UPDATE.value.format(
            radiator_name=radiator_name, current_valve_setting=current_valve_setting
        )

        if pyip.inputYesNo(change_prompt) == SCons.AGREE.value:
            return pyip.inputStr(f"\nEnter new valve setting for '{radiator_name}': ")

        return None

    def _confirm_valve_update(
        self, current_valve_setting: str, new_valve_setting: str, new_row: int
    ) -> bool:
        """Asks the user for confirmation before updating the valve setting.

        Args:
            current_valve_setting: The existing valve setting.
            new_valve_setting: The new valve setting entered by the user.
            new_row: The worksheet row index to write the new data in.

        Returns:
            True if the user confirms the change; False otherwise.
        """
        print(f"\nCurrent value: {current_valve_setting}. New value: {new_valve_setting}")

        if pyip.inputYesNo("Apply this change? (yes/no)\n") == SCons.AGREE.value:
            self.worksheet.cell(row=new_row, column=SNums.EIGHT.value, value=new_valve_setting)
            return True
        return False

    def _keep_existing_valve_setting(self, new_row: int, current_valve_setting: str) -> None:
        """Writes the existing valve setting back to the sheet if not updated.

        Args:
            new_row: The worksheet row index to write the data in.
            current_valve_setting: The existing valve setting to retain.
        """
        self.worksheet.cell(row=new_row, column=SNums.EIGHT.value, value=current_valve_setting)

    def _get_raw_readings(self) -> None:
        """Writes raw readings to the appropriate column for each radiator.

        After entry, the user is shown a recap of all raw readings and
        can make corrections.
        """
        logging.info("Prompting user for raw readings.")
        print(UserMsgs.ENTER_RAW_READINGS.value)

        radiators_owned, start_row = self._get_usage_block_info()
        self._fill_col(
            rows=range(start_row, start_row + radiators_owned),
            get_value=lambda row: pyip.inputInt(
                f"Raw Reading for {self.worksheet.cell(row=row, column=SNums.TWO.value).value}: "
            ),
            col=SNums.FIVE.value,
        )

        self._recap_and_edit_column("raw reading", SNums.FIVE.value, self._get_raw_readings)

    def _get_actual_values(self) -> None:
        """Writes the actual value for each radiator."""
        logging.info("Fillling usage file with the actual values.")
        radiators_owned, start_row = self._get_usage_block_info()
        self._fill_col(
            rows=range(start_row, start_row + radiators_owned),
            get_value=self._compute_actual_value,
            col=SNums.SIX.value,
        )

    def _compute_actual_value(self, row: int) -> float | int:
        """Computes the actual value (Coefficient * Raw Reading) for each row.

        Args:
            row: The worksheet row index.

        Returns:
            The computed actual value as a float.
        """
        coefficient = cast(float | int, self.worksheet.cell(row=row, column=SNums.FOUR.value).value)
        raw_reading = cast(int, self.worksheet.cell(row=row, column=SNums.FIVE.value).value)
        return coefficient * raw_reading

    def _get_total(self) -> None:
        """Gets the sum of actual values for all radiators.

        Sums the 'Actual Value' column for the current data block and
        writes the total to the worksheet.
        """
        logging.info('Writing the total to the usage file.')
        radiators_owned, start_row = self._get_usage_block_info()
        total_value = sum(
            cast(float | int, self.worksheet.cell(row=row, column=SNums.SIX.value).value)
            for row in range(start_row, start_row + radiators_owned)
        )
        self.worksheet.cell(
            row=start_row + radiators_owned,  # Write it below the last radiator's data row.
            column=SNums.SEVEN.value,
            value=total_value,
        )

    def _get_notes(self) -> None:
        """Prompts the user to add a note for the radiators block."""
        write_note = "\nWould you like to add a note for today's session? (yes/no)\n"

        if pyip.inputYesNo(write_note) == SCons.AGREE.value:
            note = pyip.inputStr("\nEnter your note:\n")
        else:
            note = "No additional notes."

        self._merge_valve_cells()
        self._insert_note(note)

    def _merge_valve_cells(self) -> None:
        """Merges the note column cells for the current data block."""
        radiators_owned, start_row = self._get_usage_block_info()
        self.worksheet.merge_cells(
            start_row=start_row,
            end_row=start_row + radiators_owned - SNums.ONE.value,
            start_column=SNums.NINE.value,
            end_column=SNums.NINE.value,
        )

    def _insert_note(self, note: str) -> None:
        """Writes the note at the top-left of the merged cell with wrapping.

        Args:
            note: The note to write.
        """
        logging.info('Adding a note to the usage file.')
        cell = self.worksheet.cell(
            row=self.config_data[CKeys.START_ROW.value], column=SNums.NINE.value, value=note
        )
        cell.alignment = Alignment(
            wrap_text=True,
            vertical=UsMeta.NOTE_VERTICAL_ALIGNMENT.value,
            horizontal=UsMeta.NOTE_HORIZONTAL_ALIGNMENT.value,
        )

    def _add_blank_lines(self) -> None:
        """Adds three blank lines to the worksheet for improved readability."""
        logging.info("Adding three blank lines to the usage file.")
        for _ in range(SNums.THREE.value):
            empty_row = self.worksheet.max_row + SNums.ONE.value
            self.worksheet.cell(row=empty_row, column=SNums.ONE.value, value=SCons.EMPTY_STR.value)

    def save_usage(self) -> None:
        """Saves the usage workbook to the specified file path."""
        self.excel_shared.save_workbook(self.file_path, Labels.USAGE.value)

    def open_usage(self) -> None:
        """Prompts the user to open the usage file.

        If the user agrees, it opens the file with the default app.
        """
        open_prompt = UserMsgs.OPEN_USAGE.value

        if pyip.inputYesNo(open_prompt) == SCons.AGREE.value:
            self._open_based_on_system()
            logging.info("Opening the Excel usage file for the user to review it.")

    def _open_based_on_system(self) -> None:
        """Opens the usage file with the default application.

        The opening process depends on the user's operating system.
        """
        if sys.platform.startswith(OsPlatforms.MACOS.value):
            os.system(f'{OsPlatforms.MACOS_OPEN.value} "{self.file_path}"')
        elif sys.platform.startswith(OsPlatforms.WINDOWS.value):
            subprocess.run(
                [OsPlatforms.WINDOWS_OPEN.value, self.file_path], shell=True, check=False
            )
        elif sys.platform.startswith(OsPlatforms.LINUX.value):
            os.system(f'{OsPlatforms.LINUX_OPEN.value} "{self.file_path}"')
        else:
            print("Automatic file opening is not supported on this OS.")


# endregion.
